import openpyxl

from src import Participant


def write_to_target_sheet(filepath, data: Participant):
    """
    Write participant data to target file
    :param filepath: str -> absolute file path of target file
    :param data: Participant -> participant data object
    """
    # get existing variable data sheet
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook['Variables']
    last_row = get_last_row(worksheet)
    worksheet.cell(row=last_row+1, column=1).value = data.participant
    worksheet.cell(row=last_row+1, column=5).value = data.digit_span_fwd
    worksheet.cell(row=last_row+1, column=6).value = data.digit_span_bk
    worksheet.cell(row=last_row+1, column=7).value = data.nwr_n
    worksheet.cell(row=last_row+1, column=8).value = data.nwr_e
    worksheet.cell(row=last_row+1, column=9).value = data.nwr_dc
    worksheet.cell(row=last_row+1, column=10).value = data.celf_bin
    worksheet.cell(row=last_row+1, column=11).value = data.celf_3pt
    worksheet.cell(row=last_row+1, column=12).value = data.lex_tale
    worksheet.cell(row=last_row+1, column=13).value = data.poll_sen_rep_bin
    worksheet.cell(row=last_row+1, column=14).value = data.poll_sen_rep_3pt
    worksheet.cell(row=last_row+1, column=15).value = data.poll_sen_rep_srt
    worksheet.cell(row=last_row+1, column=16).value = data.poll_sen_rep_lng
    worksheet.cell(row=last_row+1, column=17).value = data.poll_sen_rep_adj
    worksheet.cell(row=last_row+1, column=18).value = data.poll_sen_rep_arg
    worksheet.cell(row=last_row+1, column=19).value = data.poll_sen_rep_sadj
    worksheet.cell(row=last_row+1, column=20).value = data.poll_sen_rep_ladj
    worksheet.cell(row=last_row+1, column=21).value = data.poll_sen_rep_sarg
    worksheet.cell(row=last_row+1, column=22).value = data.poll_sen_rep_larg
    workbook.save(filepath)


def get_last_row(sheet_object):
    """
    Get the last populated row from a spreadsheet
    :param sheet_object: openpyxl worksheet object
    :return: int -> index of the last row
    """
    max_row_index = sheet_object.max_row
    while max_row_index > 1:
        if sheet_object.cell(max_row_index, column=1).value is None:
            max_row_index -= 1
        else:
            break
    return max_row_index
