import csv
import openpyxl
from dataclasses import dataclass, field
from typing import Any, List


@dataclass
class DataVariable:
    """Data class for each variable imported from source"""
    variable: str
    sheet: str
    data_cell: str
    data_type: str
    destination_col: int
    cell_value: Any = field(default=None, init=False)

    def __init__(self, item):
        self.variable = item[0]
        self.sheet = item[1]
        self.data_cell = item[2]
        self.data_type = item[3]
        self.destination_col = int(item[4])  # needs to be an integer


def read_variables_csv(filepath: str):
    """
    Fetch variable info, variable names, cells, data type, etc.
    converts csv data into dataclass to enable data extraction from excel source
    :param filepath: str -> absolute file path of csv file enumerating variables
    :return: variables: list[DataVariable] -> contains all the definitions
                        required to fetch and store data
    """
    variables = []
    with open(filepath) as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader, None)
        for row in csv_reader:
            variable = DataVariable(row)
            variables.append(variable)
    return variables


def fetch_data(filepath: str, variables: List[DataVariable]):
    """
    Fetch data from source excel worksheets
    :param filepath: str -> absolute file path of source excel sheet
    :param variables: list[DataVariable] -> data to extract from file
    :return: variables: list[DataVariable] -> with source cell values
    """
    workbook = openpyxl.load_workbook(filepath, data_only=True)

    for variable in variables:
        variable.cell_value = workbook[variable.sheet][variable.data_cell].value

    return variables


def get_last_row(sheet_object):
    """
    Get the last populated row from a spreadsheet
    :param sheet_object: openpyxl worksheet
    :return: int, index of the last row
    """
    max_row_index = sheet_object.max_row
    while max_row_index > 1:
        if sheet_object.cell(max_row_index, column=1).value is None:
            max_row_index -= 1
        else:
            break
    return max_row_index


def write_to_target_sheet(filepath: str, target_sheet: str, data:  List[DataVariable]):
    """
    Write data to target excel file after data extraction
    :param filepath: str -> absolute file path of target excel sheet
    :param target_sheet: str -> excel worksheet that you need to fill
    :param data: list[DataVariable] -> all the variable data, source, target info
    :return:
    """
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook[target_sheet]
    last_row = get_last_row(worksheet)
    for entry in data:
        worksheet.cell(
            row=last_row+1,
            column=entry.destination_col
        ).value = entry.cell_value
    workbook.save(filepath)


def extract(vars_file: str, source_file: str, target_file: str, target_sheet: str):
    """
    Extract data from source file and write to target file
    :param vars_file: str -> absolute path to variable definition csv
    :param source_file: str -> absolute path to source data file
    :param target_file: str -> absolute path to target data file
    :param target_sheet: str -> name of worksheet in target excel workbook
    """
    variables = read_variables_csv(vars_file)
    data = fetch_data(source_file, variables)
    write_to_target_sheet(target_file, target_sheet, data)


if __name__ == '__main__':
    my_vars_file = 'D:/Code/python_projects/hannah/tmp/variables.csv'
    my_source_file = 'D:/Code/python_projects/hannah/tmp/SampleDataSheet.xlsx'
    my_target_file = 'D:/Code/python_projects/hannah/tmp/Hebert_Data Variables.xlsx'

    extract(my_vars_file, my_source_file, my_target_file, 'Variables')
